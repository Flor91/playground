import openpyxl, os
from openpyxl.cell import get_column_letter

path = os.getcwd().split(os.sep)[:-1]
filePath = os.sep.join(path + ['out', 'me.xlsx'])
wb = openpyxl.load_workbook(filePath)


def update_data(data, day):
    """Update Excel with data extracted from different sources"""
    sheet = get_sheet(day.strftime("%B"))
    column = get_column(sheet, day)

    inv_values = {sheet.cell(row=i, column=1).value: i for i in range(2, sheet.max_row + 1)}

    update_food(data[0], sheet, inv_values, column)
    update_habits(data[1], sheet, inv_values, column)
    update_motivation(data[2], sheet, inv_values, column)

    print_date(sheet, column)
    wb.save(filePath)


def get_sheet(month):
    sheet_names = wb.get_sheet_names()

    if month not in sheet_names:
        sheet = wb.create_sheet(index=1, title=month)
        # TODO: Set up new sheet
    else:
        sheet = wb.get_sheet_by_name(month)
    return sheet


def get_column(sheet, day):
    for d in range(2, sheet.max_column + 1):
        if day.__eq__(sheet.cell(row=1, column=d).value):
            print("Column " + str(d))
            return d




def print_date(sheet, column):
    values = {i: sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)}

    col = get_column_letter(column)

    for rowOfCellObjects in sheet[col + '2':col + str(sheet.max_row)]:
        for cellObj in rowOfCellObjects:
            print(values[int(cellObj.row)], " - ", cellObj.value)


def update_cell(sheet, row, column, v):
    sheet.cell(row=row, column=column + 1).value = v


def update_food(data, sheet, inv_values, column):
    weight = data[1]
    update_cell(sheet, inv_values['weight'], column, str(weight))

    mFP = data[0]

    totals = mFP.totals
    for k, v in totals.items():
        update_cell(sheet, inv_values[k], column, v)

    meals = mFP.meals
    for meal in meals:
        name = meal.name
        entries = '; '.join([x.name for x in meal.entries])
        update_cell(sheet, inv_values[name], column, entries)


def update_habits(habits, sheet, inv_values, column):
    update_cell(sheet, inv_values['habits'], column, ', '.join(set(habits)))


def update_motivation(quote, sheet, inv_values, column):
    update_cell(sheet, inv_values['motivation'], column, quote)